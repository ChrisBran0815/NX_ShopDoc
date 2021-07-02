import sys
from openpyxl import Workbook
import NXOpen
import NXOpen.CAM

theSession = NXOpen.Session.GetSession()
workPart = theSession.Parts.Work
theUI = NXOpen.UI.GetUI()
tools = []

def main():
    rootGroup = workPart.CAMSetup.GetRoot(NXOpen.CAM.CAMSetup.View.ProgramOrder)
    obj = NXOpen.CAM.NCGroup.GetMembers(rootGroup)
    parsing_group(obj)
    wb = Workbook()

    ws =  wb.active
    ws.title = "Changed Sheet"

    ws = wb.worksheets[0]
    for i , tool in enumerate(tools):
        ws[f'A{i+1}'] = tool
        #lw(tool)
    wb.save(filename = r'C:\Temp\sample_book.xlsx')

def parsing_group(obj):
    for tagged in obj:
        if not tagged == 'NONE':
            if workPart.CAMSetup.IsGroup(tagged):
                group = NXOpen.CAM.NCGroup.GetMembers(tagged)
                parsing_group(group)
            else:
                add_tool(tagged)

def add_tool(operation):
    cutting_tool = operation.GetParent(NXOpen.CAM.CAMSetup.View.MachineTool)
    cutting_tool = cutting_tool.Name
    if not cutting_tool in tools:
        tools.append(cutting_tool)
    pass

def notinuse():
    num = theUI.SelectionManager.GetNumSelectedObjects()
    objects1 = [NXOpen.CAM.CAMObject.Null] * num

    for i in range(num):
            
        objects1[i] = theUI.SelectionManager.GetSelectedTaggedObject(i)        
        name_op = objects1[i].Name
        cutting_tool = objects1[i].GetParent(NXOpen.CAM.CAMSetup.View.MachineTool)
        cutting_tool_list = cutting_tool.Name


    # Create a Workbook
    wb = Workbook()

    ws =  wb.active
    ws.title = "Changed Sheet"

    ws = wb.worksheets[0]
    ws['A1'] = name_op
    ws['A2'] = cutting_tool_list

    wb.save(filename = r'C:\Temp\sample_book.xlsx')

if __name__ == '__main__':
    main()
